"""Excel I/O for preload vs postload comparison.

Comparison files can reach 200k rows, so reads iterate row-by-row
(openpyxl read_only) and the annotated report is streamed out with
xlsxwriter in constant-memory mode instead of building a workbook in RAM.
"""

import io
from datetime import date, datetime, time
from typing import Iterator

import xlsxwriter
from openpyxl import load_workbook

MAX_ROWS = 200_000
FAILURE_DETAIL_COLUMN = "Comparison_Failure_Detail"
RED_FILL = "#FFC7CE"  # same red the validation report uses for failing cells
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DATE_FORMAT = "yyyy-mm-dd"


def is_xlsx(filename: str | None) -> bool:
    return bool(filename) and filename.lower().endswith(".xlsx")


def read_header(file_bytes: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        header_row = ()
    finally:
        wb.close()
    return [str(h).strip() if h is not None else "" for h in header_row]


def count_data_rows(file_bytes: bytes) -> int | None:
    """Rows below the header taken from the declared sheet dimension, or None
    when the workbook does not declare one. Cheap enough for an upload guard."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        max_row = wb.active.max_row
    except Exception:
        return None
    finally:
        wb.close()
    return max(0, max_row - 1) if max_row else None


def iter_data_rows(file_bytes: bytes) -> Iterator[tuple[int, tuple]]:
    """Yields (excel_row_number, values) for every non-blank row below the header."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    try:
        excel_row = 1
        emitted = 0
        for values in ws.iter_rows(min_row=2, values_only=True):
            excel_row += 1
            if not values or all(v is None for v in values):
                continue
            emitted += 1
            if emitted > MAX_ROWS:
                raise ValueError(
                    f"File has more than {MAX_ROWS:,} rows, which is the comparison limit"
                )
            yield excel_row, values
    finally:
        wb.close()


class AnnotatedResultWriter:
    """Streams a preload-shaped workbook: the original columns plus a failure
    detail column, with mismatching preload cells filled red."""

    def __init__(self, header: list[str]):
        self._buffer = io.BytesIO()
        self._workbook = xlsxwriter.Workbook(
            self._buffer,
            {"constant_memory": True, "default_date_format": DATE_FORMAT},
        )
        self._sheet = self._workbook.add_worksheet()
        self._red = self._workbook.add_format({"bg_color": RED_FILL})
        self._red_date = self._workbook.add_format(
            {"bg_color": RED_FILL, "num_format": DATE_FORMAT}
        )
        self._column_count = len(header)
        self._row = 0
        for col, name in enumerate([*header, FAILURE_DETAIL_COLUMN]):
            self._sheet.write_string(0, col, name)

    def write_row(self, values: tuple, red_columns: set[int], detail: str) -> None:
        self._row += 1
        for col in range(self._column_count):
            value = values[col] if col < len(values) else None
            self._write_cell(self._row, col, value, col in red_columns)
        if detail:
            self._sheet.write_string(self._row, self._column_count, detail)

    def _write_cell(self, row: int, col: int, value, red: bool) -> None:
        # Typed writes skip xlsxwriter's per-cell token detection, which is the
        # difference between minutes and seconds on a 200k-row report.
        cell_format = self._red if red else None
        if value is None:
            if red:
                self._sheet.write_blank(row, col, None, cell_format)
            return
        if isinstance(value, str):
            self._sheet.write_string(row, col, value, cell_format)
            return
        if isinstance(value, bool):
            self._sheet.write_boolean(row, col, value, cell_format)
            return
        if isinstance(value, (int, float)):
            self._sheet.write_number(row, col, value, cell_format)
            return
        if isinstance(value, (datetime, date, time)):
            self._sheet.write_datetime(row, col, value, self._red_date if red else None)
            return
        self._sheet.write(row, col, value, cell_format)

    def finish(self) -> bytes:
        self._workbook.close()
        return self._buffer.getvalue()
