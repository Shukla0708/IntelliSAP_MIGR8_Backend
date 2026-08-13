import csv
import io
from openpyxl import load_workbook

SOURCE_FIELD_ALIASES = {"field", "field name", "fieldname", "source field", "source field name"}
DESCRIPTION_ALIASES = {"description", "field description", "desc"}
TABLE_ALIASES = {"sap table", "table", "target table"}
TARGET_FIELD_ALIASES = {"sap field", "field", "fieldname", "target field", "target field name"}
KEY_FIELD_ALIASES = {"key field", "is key", "key", "key field flag", "is key field"}
DATATYPE_ALIASES = {"datatype", "data type", "type"}
TABLE_DESCRIPTION_ALIASES = {"table description", "sap table description", "table desc"}

TRUE_VALUES = {"y", "yes", "x", "true", "1"}


def _to_bool(value: str | None) -> bool:
    if value is None:
        return False
    return value.strip().lower() in TRUE_VALUES


def _normalize(header: str) -> str:
    return " ".join(str(header).strip().lower().split())


def _read_rows(file_bytes: bytes, filename: str) -> tuple[list[str], list[list]]:
    """Returns (header_row, data_rows) for a .csv or .xlsx/.xls upload."""
    is_csv = filename.lower().endswith(".csv")
    if is_csv:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return [], []
    header = [_normalize(h) if h is not None else "" for h in rows[0]]
    return header, rows[1:]


def _column_index(header: list[str], aliases: set[str], required_label: str) -> int:
    for idx, name in enumerate(header):
        if name in aliases:
            return idx
    raise ValueError(
        f"Could not find a '{required_label}' column. Expected one of: {', '.join(sorted(aliases))}"
    )


def _cell(row: list, idx: int) -> str | None:
    if idx >= len(row) or row[idx] is None:
        return None
    value = str(row[idx]).strip()
    return value or None


def parse_source_fields(file_bytes: bytes, filename: str) -> list[dict]:
    header, rows = _read_rows(file_bytes, filename)
    if not rows:
        return []

    field_idx = _column_index(header, SOURCE_FIELD_ALIASES, "Field Name")
    desc_idx = _column_index(header, DESCRIPTION_ALIASES, "Description")
    key_idx = _column_index(header, KEY_FIELD_ALIASES, "Key Field")
    datatype_idx = _column_index(header, DATATYPE_ALIASES, "Datatype")

    fields = []
    for row in rows:
        field_name = _cell(row, field_idx)
        if not field_name:
            continue
        fields.append({
            "field_name": field_name,
            "description": _cell(row, desc_idx),
            "key_field": _to_bool(_cell(row, key_idx)),
            "datatype": _cell(row, datatype_idx),
        })
    return fields


def parse_target_fields(file_bytes: bytes, filename: str) -> list[dict]:
    header, rows = _read_rows(file_bytes, filename)
    if not rows:
        return []

    table_idx = _column_index(header, TABLE_ALIASES, "SAP Table")
    field_idx = _column_index(header, TARGET_FIELD_ALIASES, "SAP Field")
    desc_idx = _column_index(header, DESCRIPTION_ALIASES, "Description")
    table_desc_idx = _column_index(header, TABLE_DESCRIPTION_ALIASES, "Table Description")
    datatype_idx = _column_index(header, DATATYPE_ALIASES, "Datatype")

    fields = []
    for row in rows:
        sap_table = _cell(row, table_idx)
        sap_field = _cell(row, field_idx)
        if not sap_table or not sap_field:
            continue
        fields.append({
            "sap_table": sap_table,
            "sap_field": sap_field,
            "description": _cell(row, desc_idx),
            "table_description": _cell(row, table_desc_idx),
            "datatype": _cell(row, datatype_idx),
        })
    return fields
