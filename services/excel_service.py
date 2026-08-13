import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from services.rules_engine import is_empty_raw, normalize_key, validate_cell

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MAX_STORED_EXCEPTIONS = 20
MAX_EXCEPTIONS_PER_TYPE = 5


def extract_headers(file_bytes: bytes) -> list[str]:
    """Read only the header row — used right after upload to populate the
    'Validation Rules Configuration' UI with real column names."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    return [str(h).strip() for h in header_row if h is not None]


def run_validation(file_bytes: bytes, field_configs: list[dict]):
    """
    field_configs: one dict per field, matching ValidationField columns.

    If two or more fields have flag_key, uniqueness is checked on the
    combined values (composite key), not independently per column.

    Returns (annotated_workbook_bytes, stats_dict, exceptions_list).
    Exceptions are a mixed sample: at most MAX_EXCEPTIONS_PER_TYPE rows
    per error type, and at most MAX_STORED_EXCEPTIONS overall.
    Output keeps the same layout/format as the source file, with:
      - failing cells filled red
      - one appended 'Validation_Failure_Reason' column containing the
        combined reasons ("<Field>: <reason>; ...") for every failing
        cell in that row.
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    name_to_col = _header_index(header)
    reason_col_idx = len(header)
    ws.cell(row=1, column=reason_col_idx + 1, value="Validation_Failure_Reason")

    cfg_by_field = {c["field_name"]: c for c in field_configs}
    key_field_names = [c["field_name"] for c in field_configs if c["flag_key"]]
    # Two or more key fields = one composite unique constraint, not per-column.
    composite_keys = len(key_field_names) >= 2
    seen_keys_by_field = (
        {} if composite_keys
        else {c["field_name"]: set() for c in field_configs if c["flag_key"]}
    )
    seen_composites: dict[tuple[str, ...], int] = {}

    total_rows = valid_rows = invalid_rows = total_errors = critical_errors = 0
    errors_by_field: dict[str, int] = {}
    errors_by_type: dict[str, int] = {}
    exceptions: list[dict] = []
    stored_rows_by_type: dict[str, set[int]] = {}

    for row in ws.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue  # skip fully blank trailing rows

        total_rows += 1
        row_reasons: list[str] = []
        row_has_error = False

        for field_name, cfg in cfg_by_field.items():
            col_idx = _col_index(name_to_col, field_name)
            if col_idx is None:
                continue
            cell = row[col_idx]
            seen_keys = None if composite_keys else seen_keys_by_field.get(field_name, set())
            reasons = validate_cell(cell.value, cfg, seen_keys)

            if reasons:
                row_has_error = True
                cell.fill = RED_FILL
                is_critical = cfg["flag_mandatory"] or cfg["flag_key"]

                for reason in reasons:
                    row_reasons.append(f"{field_name}: {reason}")
                    total_errors += 1
                    errors_by_field[field_name] = errors_by_field.get(field_name, 0) + 1
                    bucket = reason.split(" ")[0]
                    errors_by_type[bucket] = errors_by_type.get(bucket, 0) + 1
                    if is_critical:
                        critical_errors += 1

                    _try_store_exception(exceptions, stored_rows_by_type, {
                        "row_number": row[0].row,
                        "field_name": field_name,
                        "actual_value": str(cell.value),
                        "expected_value": _expected_label(cfg),
                        "error_type": reason,
                        "severity": "error" if is_critical else "warning",
                    })

        if composite_keys:
            dup_count = _flag_duplicate_composite(
                row, key_field_names, name_to_col, seen_composites,
                row_reasons, exceptions, stored_rows_by_type,
                errors_by_field, errors_by_type,
            )
            if dup_count:
                row_has_error = True
                total_errors += dup_count
                critical_errors += dup_count

        if row_has_error:
            invalid_rows += 1
            ws.cell(row=row[0].row, column=reason_col_idx + 1, value="; ".join(row_reasons))
        else:
            valid_rows += 1

    out = io.BytesIO()
    wb.save(out)
    wb.close()

    health_score = round((valid_rows / total_rows) * 100, 2) if total_rows else 100.0

    # errorsByType is rendered as a donut chart on the frontend and expects a
    # hex color per slice plus a percentage-of-total value (not a raw count).
    PALETTE = ["#004da4", "#6063ee", "#8a3500", "#c2c6d5", "#0f9d58", "#d93025"]
    type_total = sum(errors_by_type.values()) or 1
    errors_by_type_chart = [
        {
            "label": label,
            "value": round((count / type_total) * 100),
            "color": PALETTE[i % len(PALETTE)],
        }
        for i, (label, count) in enumerate(errors_by_type.items())
    ]

    stats = {
        "total_records": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "total_errors": total_errors,
        "critical_errors": critical_errors,
        "health_score": health_score,
        "errors_by_field": [{"field": k, "count": v} for k, v in errors_by_field.items()],
        "errors_by_type": errors_by_type_chart,
    }
    return out.getvalue(), stats, exceptions


def _header_index(header: list) -> dict[str, int]:
    """Map header names to 0-based indexes (exact and lowercase)."""
    mapping: dict[str, int] = {}
    for idx, name in enumerate(header):
        if name is None:
            continue
        label = str(name).strip()
        if not label:
            continue
        mapping[label] = idx
        mapping[label.lower()] = idx
    return mapping


def _col_index(name_to_col: dict[str, int], field_name: str) -> int | None:
    if field_name in name_to_col:
        return name_to_col[field_name]
    return name_to_col.get(field_name.lower())


def _exception_type_key(reason: str) -> str:
    """Group composite-key duplicates as one type regardless of first-row text."""
    if reason.startswith("Duplicate composite key value"):
        return "Duplicate composite key value"
    return reason


def _try_store_exception(
    exceptions: list[dict],
    stored_rows_by_type: dict[str, set[int]],
    item: dict,
) -> None:
    """Keep a mixed sample: up to N rows per error type, and an overall cap."""
    if len(exceptions) >= MAX_STORED_EXCEPTIONS:
        return
    key = _exception_type_key(item["error_type"])
    rows = stored_rows_by_type.setdefault(key, set())
    row_num = item["row_number"]
    if row_num not in rows and len(rows) >= MAX_EXCEPTIONS_PER_TYPE:
        return
    exceptions.append(item)
    rows.add(row_num)


def _row_number(row) -> int:
    for cell in row:
        if cell is not None:
            return cell.row
    return 0


def _flag_duplicate_composite(
    row,
    key_field_names: list[str],
    name_to_col: dict[str, int],
    seen_composites: dict[tuple[str, ...], int],
    row_reasons: list[str],
    exceptions: list[dict],
    stored_rows_by_type: dict[str, set[int]],
    errors_by_field: dict[str, int],
    errors_by_type: dict[str, int],
) -> int:
    """Flag a duplicate composite key. Returns the number of new errors (0 if unique)."""
    parts: list[str] = []
    key_cells: list[tuple[str, object, str]] = []
    for field_name in key_field_names:
        col_idx = _col_index(name_to_col, field_name)
        if col_idx is None or col_idx >= len(row):
            return 0
        cell = row[col_idx]
        raw = normalize_key(cell.value)
        parts.append(raw)
        key_cells.append((field_name, cell, raw))

    # Incomplete keys are already flagged per-cell; uniqueness needs every part.
    if any(is_empty_raw(p) for p in parts):
        return 0

    composite = tuple(parts)
    row_num = _row_number(row)
    first_row = seen_composites.get(composite)
    if first_row is None:
        seen_composites[composite] = row_num
        return 0

    reason = f"Duplicate composite key value (same as row {first_row})"
    label = " + ".join(key_field_names)
    for field_name, cell, raw in key_cells:
        cell.fill = RED_FILL
        row_reasons.append(f"{field_name}: {reason}")
        errors_by_field[field_name] = errors_by_field.get(field_name, 0) + 1
        errors_by_type["Duplicate"] = errors_by_type.get("Duplicate", 0) + 1
        _try_store_exception(exceptions, stored_rows_by_type, {
            "row_number": row_num,
            "field_name": field_name,
            "actual_value": raw,
            "expected_value": f"Unique composite ({label})",
            "error_type": reason,
            "severity": "error",
        })
    return len(key_cells)


def _expected_label(cfg: dict) -> str:
    if cfg["flag_key"]:
        return "Non-empty unique key"
    if cfg["flag_email"]:
        return "Valid email format"
    if cfg["flag_mobile"]:
        return "Valid mobile number"
    if cfg["flag_date"]:
        return "Valid date"
    if cfg["max_length"]:
        return f"Max length {cfg['max_length']}"
    return cfg["data_type"]
