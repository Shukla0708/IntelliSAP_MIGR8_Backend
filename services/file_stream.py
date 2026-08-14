"""Stream CSV/XLSX without loading a full openpyxl workbook into memory."""
from __future__ import annotations

import csv
import io
import logging
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

CHUNK_SIZE = 25_000


def sniff_kind(filename: str) -> str:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def extract_headers(file_bytes: bytes, filename: str = "source.xlsx") -> list[str]:
    kind = sniff_kind(filename)
    if kind == "csv":
        text = file_bytes[:16384].decode("utf-8-sig", errors="replace")
        first = next((line for line in text.splitlines() if line.strip()), "")
        if not first:
            return []
        header = next(csv.reader([first]))
        return [str(h).strip() for h in header if h is not None and str(h).strip()]
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    return [str(h).strip() for h in header_row if h is not None and str(h).strip()]


def extract_headers_from_path(path: Path, filename: str) -> list[str]:
    kind = sniff_kind(filename)
    if kind == "csv":
        with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as handle:
            reader = csv.reader(handle)
            for row in reader:
                return [str(h).strip() for h in row if h is not None and str(h).strip()]
        return []
    try:
        import polars as pl

        frame = pl.read_excel(
            str(path), engine="calamine", infer_schema_length=0, n_rows=1,
        )
        return [str(name).strip() for name in frame.columns if str(name).strip()]
    except Exception:
        logger.exception("calamine header read failed; falling back to openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        wb.close()
        return [str(h).strip() for h in header_row if h is not None and str(h).strip()]


def iter_data_rows(path: Path, filename: str) -> Iterator[tuple[int, list]]:
    """Yield (1-based Excel row number, values) for data rows. Header is row 1."""
    if sniff_kind(filename) == "csv":
        yield from _iter_csv(path)
        return
    try:
        yield from _iter_xlsx_calamine(path)
    except Exception:
        logger.exception("calamine/polars XLSX read failed; falling back to openpyxl read_only")
        yield from _iter_xlsx_openpyxl(path)


def _iter_csv(path: Path) -> Iterator[tuple[int, list]]:
    with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as handle:
        reader = csv.reader(handle)
        header = next(reader, None)
        if header is None:
            return
        excel_row = 1
        for raw in reader:
            excel_row += 1
            if not any(cell not in (None, "") for cell in raw):
                continue
            yield excel_row, list(raw)


def _iter_xlsx_calamine(path: Path) -> Iterator[tuple[int, list]]:
    import polars as pl

    frame = pl.read_excel(str(path), engine="calamine", infer_schema_length=0)
    excel_row = 1
    for raw in frame.iter_rows():
        excel_row += 1
        values = ["" if cell is None else cell for cell in raw]
        if not any(cell not in (None, "") for cell in values):
            continue
        yield excel_row, values


def _iter_xlsx_openpyxl(path: Path) -> Iterator[tuple[int, list]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    try:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return
        excel_row = 1
        for raw in rows:
            excel_row += 1
            if not raw or all(cell is None or cell == "" for cell in raw):
                continue
            yield excel_row, ["" if cell is None else cell for cell in raw]
    finally:
        wb.close()
