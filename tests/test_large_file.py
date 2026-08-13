"""CSV streaming validation plus preload/postload comparison."""
import csv
import io

from openpyxl import load_workbook

from services.comparison_engine import run_comparison
from services.excel_service import run_validation
from tests.test_composite_keys import _cfg, _xlsx


def _csv(headers, rows) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def test_csv_composite_duplicate():
    data = _csv(
        ["CustomerID", "OrderID", "Name"],
        [
            [1, "A", "Ann"],
            [1, "B", "Bob"],
            [1, "A", "Dee"],
        ],
    )
    result_bytes, stats, exceptions = run_validation(
        data,
        [_cfg("CustomerID", True), _cfg("OrderID", True), _cfg("Name")],
        filename="source.csv",
    )
    assert stats["invalid_rows"] == 1
    assert stats["valid_rows"] == 2
    wb = load_workbook(io.BytesIO(result_bytes))
    ws = wb.active
    assert ws.cell(row=1, column=4).value == "Validation_Failure_Reason"
    assert "Duplicate composite key" in (ws.cell(row=4, column=4).value or "")


def test_comparison_detects_mismatch_and_drop(tmp_path):
    preload = tmp_path / "pre.csv"
    postload = tmp_path / "post.csv"
    preload.write_text(
        "ID,EMAIL,NAME\n100,a@old.com,Ann\n200,b@ok.com,Bob\n300,c@x.com,Cal\n",
        encoding="utf-8",
    )
    postload.write_text(
        "ID,EMAIL,NAME\n100,a@new.com,Ann\n200,b@ok.com,Bob\n",
        encoding="utf-8",
    )
    result_bytes, stats, exceptions = run_comparison(
        preload,
        "pre.csv",
        postload,
        "post.csv",
        join_keys=[("ID", "ID")],
        compare_fields=[("EMAIL", "EMAIL"), ("NAME", "NAME")],
    )
    assert stats["matched_records"] == 1
    assert stats["different_count"] == 1
    assert stats["missing_count"] == 1
    types = {e["difference_type"] for e in exceptions}
    assert "VALUE_MISMATCH" in types
    assert "DROPPED_RECORD" in types
    assert result_bytes[:2] == b"PK"
