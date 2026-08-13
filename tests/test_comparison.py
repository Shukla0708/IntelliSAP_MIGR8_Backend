"""Tests for preload vs postload comparison runs."""
import io
import uuid

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from main import app
from db.database import SessionLocal
from db.models import ComparisonRun, FinalMapping, Mapping, User
from auth import hash_password, create_access_token
from services import s3_service
from services.comparison_file_service import FAILURE_DETAIL_COLUMN

HEADER = ["CUSTOMER_ID", "COMPANY_CODE", "EMAIL", "POSTAL_CODE"]
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _make_user(db) -> tuple[User, dict]:
    suffix = uuid.uuid4().hex[:12]
    user = User(
        full_name="Comparison Tester",
        email=f"comparison-{suffix}@example.com",
        password_hash=hash_password("test-password-123"),
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user, {"Authorization": f"Bearer {create_access_token(str(user.id))}"}


@pytest.fixture
def auth_user(db):
    user, headers = _make_user(db)
    yield {"user": user, "headers": headers}
    db.delete(db.get(User, user.id) or user)
    db.commit()


def _workbook_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _create_project(client: TestClient, headers: dict) -> str:
    res = client.post(
        "/api/projects/", json={"name": f"Compare-{uuid.uuid4().hex[:8]}"}, headers=headers
    )
    assert res.status_code == 200, res.text
    return res.json()["id"]


def _create_run(client: TestClient, headers: dict, project_id: str) -> str:
    res = client.post(
        f"/api/comparisons/?project_id={project_id}",
        json={"name": f"Run {uuid.uuid4().hex[:8]}"},
        headers=headers,
    )
    assert res.status_code == 200, res.text
    return res.json()["run_id"]


def _upload(client, headers, run_id, preload_rows, postload_rows, preload_name="preload.xlsx"):
    return client.post(
        f"/api/comparisons/{run_id}/upload",
        files={
            "preload_file": (preload_name, _workbook_bytes(preload_rows), XLSX),
            "postload_file": ("postload.xlsx", _workbook_bytes(postload_rows), XLSX),
        },
        headers=headers,
    )


def _add_mapping(db, project_id: str, fields: list[tuple[str, str, bool]]) -> str:
    mapping = Mapping(
        project_id=uuid.UUID(project_id), mapping_name="Test mapping", status="completed"
    )
    db.add(mapping)
    db.commit()
    db.refresh(mapping)
    for source_field, target_field, is_key in fields:
        db.add(FinalMapping(
            mapping_id=mapping.id,
            source_field=source_field,
            target_field=target_field,
            key=is_key,
        ))
    db.commit()
    return str(mapping.id)


def _by_type(discrepancies: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}
    for item in discrepancies:
        grouped.setdefault(item["differenceType"], []).append(item)
    return grouped


def test_same_name_columns_compare_without_mapping(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    uploaded = _upload(
        client, headers, run_id,
        [HEADER,
         ["100045", "1000", "john@old.com", "90210"],
         ["100082", "1000", "amy@x.com", "90210"],
         ["100119", "2000", "gone@x.com", "11111"]],
        [HEADER,
         ["100045", "1000", "john@new.com", "90210"],
         ["100082", "1000", "amy@x.com", "90210"]],
    )
    assert uploaded.status_code == 200, uploaded.text
    assert uploaded.json()["preload_fields"] == HEADER

    executed = client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)
    assert executed.status_code == 200, executed.text

    result = client.get(f"/api/comparisons/{run_id}/result", headers=headers)
    assert result.status_code == 200, result.text
    body = result.json()
    assert body["status"] == "completed"
    assert body["matchedRecords"] == 1
    assert body["differentCount"] == 1
    assert body["missingCount"] == 1
    assert body["matchRate"] == "33.3% Match Rate"

    grouped = _by_type(body["discrepancies"])
    assert grouped["VALUE_MISMATCH"][0]["field"] == "EMAIL"
    assert grouped["VALUE_MISMATCH"][0]["status"] == "warning"
    dropped = grouped["DROPPED_RECORD"][0]
    assert dropped["status"] == "error"
    assert dropped["fieldItalic"] is True
    assert "100119" in dropped["businessKey"]


def test_explicit_composite_key_without_mapping(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    # Same CUSTOMER_ID in two company codes — only the pair identifies a row.
    _upload(
        client, headers, run_id,
        [HEADER,
         ["100045", "1000", "a@x.com", "90210"],
         ["100045", "2000", "b@x.com", "90211"]],
        [HEADER,
         ["100045", "1000", "a@x.com", "90210"],
         ["100045", "2000", "changed@x.com", "90211"]],
    )

    executed = client.post(
        f"/api/comparisons/{run_id}/execute",
        json={
            "business_key_columns_preload": ["CUSTOMER_ID", "COMPANY_CODE"],
            "business_key_columns_postload": ["CUSTOMER_ID", "COMPANY_CODE"],
        },
        headers=headers,
    )
    assert executed.status_code == 200, executed.text

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    assert body["matchedRecords"] == 1
    assert body["differentCount"] == 1
    assert body["missingCount"] == 0
    mismatch = _by_type(body["discrepancies"])["VALUE_MISMATCH"][0]
    assert mismatch["businessKey"] == "CUSTOMER_ID: 100045 | COMPANY_CODE: 2000"


def test_mapping_with_single_key_field(client, auth_user, db):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    mapping_id = _add_mapping(db, project_id, [
        ("CUSTOMER_ID", "KNA1.CUSTOMER_ID", True),
        ("EMAIL", "KNA1.EMAIL", False),
    ])
    run_id = _create_run(client, headers, project_id)

    _upload(
        client, headers, run_id,
        [HEADER,
         ["100045", "1000", "john@old.com", "90210"],
         ["100082", "1000", "amy@x.com", "90210"]],
        [HEADER,
         ["100045", "9999", "john@new.com", "00000"],
         ["100082", "9999", "amy@x.com", "00000"]],
    )

    available = client.get(f"/api/comparisons/{run_id}/available-mappings", headers=headers)
    assert available.status_code == 200, available.text
    listed = next(item for item in available.json() if item["id"] == mapping_id)
    assert listed["confirmedFieldCount"] == 2
    assert listed["keyFieldCount"] == 1

    executed = client.post(
        f"/api/comparisons/{run_id}/execute", json={"mapping_id": mapping_id}, headers=headers
    )
    assert executed.status_code == 200, executed.text

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    # Only EMAIL is compared, so the differing COMPANY_CODE/POSTAL_CODE are ignored.
    assert body["matchedRecords"] == 1
    assert body["differentCount"] == 1
    assert _by_type(body["discrepancies"])["VALUE_MISMATCH"][0]["field"] == "EMAIL"


def test_mapping_with_composite_key(client, auth_user, db):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    mapping_id = _add_mapping(db, project_id, [
        ("CUSTOMER_ID", "KNA1.CUSTOMER_ID", True),
        ("COMPANY_CODE", "KNA1.COMPANY_CODE", True),
        ("EMAIL", "KNA1.EMAIL", False),
        ("POSTAL_CODE", "KNA1.POSTAL_CODE", False),
    ])
    run_id = _create_run(client, headers, project_id)

    _upload(
        client, headers, run_id,
        [HEADER,
         ["100045", "1000", "a@x.com", "90210"],
         ["100045", "2000", "b@x.com", "90211"]],
        [HEADER,
         ["100045", "1000", "a@x.com", "90210"],
         ["100045", "2000", "b@x.com", "90210"]],
    )

    executed = client.post(
        f"/api/comparisons/{run_id}/execute", json={"mapping_id": mapping_id}, headers=headers
    )
    assert executed.status_code == 200, executed.text

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    assert body["matchedRecords"] == 1
    assert body["differentCount"] == 1
    assert body["missingCount"] == 0
    mismatch = _by_type(body["discrepancies"])["VALUE_MISMATCH"][0]
    assert mismatch["field"] == "POSTAL_CODE"
    assert mismatch["businessKey"] == "CUSTOMER_ID: 100045 | COMPANY_CODE: 2000"


def test_mapping_without_key_field_returns_422(client, auth_user, db):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    mapping_id = _add_mapping(db, project_id, [("EMAIL", "KNA1.EMAIL", False)])
    run_id = _create_run(client, headers, project_id)

    _upload(
        client, headers, run_id,
        [HEADER, ["100045", "1000", "a@x.com", "90210"]],
        [HEADER, ["100045", "1000", "a@x.com", "90210"]],
    )

    executed = client.post(
        f"/api/comparisons/{run_id}/execute", json={"mapping_id": mapping_id}, headers=headers
    )
    assert executed.status_code == 422
    assert "key field" in executed.json()["detail"]


def test_format_change_is_reported_as_info(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    _upload(
        client, headers, run_id,
        [HEADER, ["100045", "1000", "john@x.com", "90210"]],
        [HEADER, ["100045", "1000", "JOHN@X.COM", "90210"]],
    )
    client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    change = _by_type(body["discrepancies"])["FORMAT_CHANGE"][0]
    assert change["status"] == "info"
    assert change["field"] == "EMAIL"


def test_zero_padded_keys_and_values_still_match(client, auth_user):
    """SAP's ALPHA conversion pads keys, which must not read as a missing record."""
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    header = ["CUSTOMER_ID", "COMPANY_CODE", "EMAIL"]
    _upload(
        client, headers, run_id,
        [header, ["100045", "1000", "a@x.com"]],
        [header, ["0000100045", "0000001000", "a@x.com"]],
    )
    client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    assert body["matchedRecords"] == 1
    assert body["differentCount"] == 0
    assert body["missingCount"] == 0
    assert body["discrepancies"] == []


def test_reformatted_numbers_and_dates_are_not_differences(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    header = ["DOC_ID", "AMOUNT", "POSTING_DATE", "CREDIT"]
    _upload(
        client, headers, run_id,
        [header, ["500001", "1000", "2024-01-05", "-250.5"]],
        [header, ["0500001", "1,000.00", "05.01.2024", "250.50-"]],
    )
    client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    assert body["matchedRecords"] == 1
    assert body["differentCount"] == 0
    assert body["discrepancies"] == []


def test_real_value_changes_survive_normalization(client, auth_user):
    """The equivalence rules must not swallow an actual change."""
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    header = ["CUSTOMER_ID", "AMOUNT", "DISCOUNT"]
    _upload(
        client, headers, run_id,
        [header, ["100045", "1000", "0"]],
        [header, ["100045", "1000.01", ""]],
    )
    client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    assert body["matchedRecords"] == 0
    assert body["differentCount"] == 1
    mismatched = {row["field"] for row in _by_type(body["discrepancies"])["VALUE_MISMATCH"]}
    assert mismatched == {"AMOUNT", "DISCOUNT"}


def test_extra_postload_record_is_reported(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    _upload(
        client, headers, run_id,
        [HEADER, ["100045", "1000", "a@x.com", "90210"]],
        [HEADER,
         ["100045", "1000", "a@x.com", "90210"],
         ["100999", "1000", "new@x.com", "11111"]],
    )
    client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    assert body["matchedRecords"] == 1
    assert body["missingCount"] == 1
    extra = _by_type(body["discrepancies"])["EXTRA_RECORD"][0]
    assert extra["status"] == "error"
    assert "100999" in extra["businessKey"]


def test_discrepancies_are_capped_at_fifty(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    preload = [HEADER] + [[f"{i:06d}", "1000", f"old{i}@x.com", "90210"] for i in range(60)]
    postload = [HEADER] + [[f"{i:06d}", "1000", f"new{i}@x.com", "90211"] for i in range(60)]
    _upload(client, headers, run_id, preload, postload)
    client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)

    body = client.get(f"/api/comparisons/{run_id}/result", headers=headers).json()
    assert body["differentCount"] == 60
    assert len(body["discrepancies"]) == 50


def test_downloaded_report_has_preload_layout(client, auth_user, db):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    _upload(
        client, headers, run_id,
        [HEADER,
         ["100045", "1000", "john@old.com", "90210"],
         ["100082", "1000", "amy@x.com", "90210"]],
        [HEADER,
         ["100045", "1000", "john@new.com", "90210"],
         ["100082", "1000", "amy@x.com", "90210"]],
    )
    client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)

    url = client.get(f"/api/comparisons/{run_id}/download-url", headers=headers)
    assert url.status_code == 200, url.text
    assert url.json()["url"]

    # Read the stored workbook directly so the test holds for either storage backend.
    run = db.get(ComparisonRun, uuid.UUID(run_id))
    report = s3_service.download_bytes(run.result_s3_key)

    sheet = load_workbook(io.BytesIO(report)).active
    rows = list(sheet.iter_rows(values_only=True))
    assert list(rows[0]) == [*HEADER, FAILURE_DETAIL_COLUMN]
    assert rows[1][0] == "100045"
    assert rows[1][-1] == "EMAIL: preload=john@old.com | postload=john@new.com"
    assert rows[2][-1] is None
    # The failing preload cell is filled red, matching the validation report.
    assert sheet.cell(row=2, column=3).fill.start_color.rgb == "FFFFC7CE"


def test_project_mapping_list_reports_key_counts(client, auth_user, db):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    mapping_id = _add_mapping(db, project_id, [
        ("CUSTOMER_ID", "KNA1.CUSTOMER_ID", True),
        ("COMPANY_CODE", "KNA1.COMPANY_CODE", True),
        ("EMAIL", "KNA1.EMAIL", False),
    ])

    listed = client.get(f"/api/mappings/?project_id={project_id}", headers=headers)
    assert listed.status_code == 200, listed.text
    match = next(item for item in listed.json() if item["mappingRunId"] == mapping_id)
    assert match["confirmedFieldCount"] == 3
    assert match["keyFieldCount"] == 2

    confirmed = client.get(f"/api/mappings/{mapping_id}/confirmed", headers=headers)
    assert confirmed.status_code == 200, confirmed.text
    keys = [field["sourceField"] for field in confirmed.json()["fields"] if field["isKey"]]
    assert keys == ["COMPANY_CODE", "CUSTOMER_ID"]


def test_upload_rejects_non_xlsx(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    res = _upload(
        client, headers, run_id,
        [HEADER, ["100045", "1000", "a@x.com", "90210"]],
        [HEADER, ["100045", "1000", "a@x.com", "90210"]],
        preload_name="preload.csv",
    )
    assert res.status_code == 400
    assert "preload.csv" in res.json()["detail"]


def test_execute_before_upload_returns_400(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    res = client.post(f"/api/comparisons/{run_id}/execute", json={}, headers=headers)
    assert res.status_code == 400


def test_duplicate_run_name_returns_409(client, auth_user):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    body = {"name": "Duplicate comparison"}

    first = client.post(f"/api/comparisons/?project_id={project_id}", json=body, headers=headers)
    second = client.post(f"/api/comparisons/?project_id={project_id}", json=body, headers=headers)
    assert first.status_code == 200, first.text
    assert second.status_code == 409


def test_list_is_scoped_to_owner(client, auth_user, db):
    headers = auth_user["headers"]
    project_id = _create_project(client, headers)
    run_id = _create_run(client, headers, project_id)

    listed = client.get("/api/comparisons/", headers=headers).json()
    match = next(row for row in listed if row["id"] == run_id)
    assert match["projectId"] == project_id
    assert match["status"] == "draft"

    other_user, other_headers = _make_user(db)
    try:
        assert all(row["id"] != run_id for row in client.get("/api/comparisons/", headers=other_headers).json())
        assert client.get(f"/api/comparisons/{run_id}/result", headers=other_headers).status_code == 404
    finally:
        db.delete(db.get(User, other_user.id) or other_user)
        db.commit()
