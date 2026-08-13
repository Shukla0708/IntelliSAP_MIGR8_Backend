"""Composite and single-key uniqueness in run_validation."""
import io

from openpyxl import Workbook, load_workbook

from services.excel_service import run_validation


def _cfg(name, key=False, mandatory=False):
    return {
        "field_name": name,
        "flag_key": key,
        "flag_mandatory": mandatory,
        "flag_null": False,
        "flag_email": False,
        "flag_mobile": False,
        "flag_date": False,
        "flag_special_chars": False,
        "case_format": None,
        "data_type": "string",
        "max_length": None,
        "decimal_length": None,
        "regex": None,
    }


def _xlsx(headers, rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_composite_allows_shared_single_parts():
    data = _xlsx(
        ["CustomerID", "OrderID", "Name"],
        [
            [1, "A", "Ann"],
            [1, "B", "Bob"],
            [2, "A", "Cal"],
        ],
    )
    _, stats, exceptions = run_validation(
        data, [_cfg("CustomerID", True), _cfg("OrderID", True), _cfg("Name")]
    )
    assert stats["invalid_rows"] == 0
    assert exceptions == []


def test_composite_flags_duplicate_pair_not_first_row():
    data = _xlsx(
        ["CustomerID", "OrderID", "Name"],
        [
            [1, "A", "Ann"],
            [1, "B", "Bob"],
            [1, "A", "Dee"],
        ],
    )
    result_bytes, stats, exceptions = run_validation(
        data, [_cfg("CustomerID", True), _cfg("OrderID", True), _cfg("Name")]
    )
    assert stats["valid_rows"] == 2
    assert stats["invalid_rows"] == 1
    assert all(e["row_number"] == 4 for e in exceptions)
    assert {e["field_name"] for e in exceptions} == {"CustomerID", "OrderID"}
    assert all("same as row 2" in e["error_type"] for e in exceptions)

    wb = load_workbook(io.BytesIO(result_bytes))
    ws = wb.active
    assert "Duplicate composite key" in (ws.cell(row=4, column=4).value or "")
    assert ws.cell(row=2, column=4).value is None


def test_composite_matches_excel_int_and_float():
    data = _xlsx(
        ["CustomerID", "OrderID"],
        [
            [1, 10],
            [1.0, 10.0],
        ],
    )
    _, stats, exceptions = run_validation(
        data, [_cfg("CustomerID", True), _cfg("OrderID", True)]
    )
    assert stats["invalid_rows"] == 1
    assert exceptions


def test_composite_skips_uniqueness_when_a_key_is_empty():
    data = _xlsx(
        ["CustomerID", "OrderID"],
        [
            [1, None],
            [1, None],
        ],
    )
    _, stats, exceptions = run_validation(
        data, [_cfg("CustomerID", True), _cfg("OrderID", True)]
    )
    assert all(e["error_type"] == "Key value is empty" for e in exceptions)
    assert not any("Duplicate" in e["error_type"] for e in exceptions)
    assert stats["invalid_rows"] == 2


def test_single_key_still_independent():
    data = _xlsx(
        ["CustomerID", "Name"],
        [
            ["1", "Ann"],
            ["1", "Bob"],
        ],
    )
    _, stats, exceptions = run_validation(
        data, [_cfg("CustomerID", True), _cfg("Name")]
    )
    assert stats["invalid_rows"] == 1
    assert any(e["error_type"] == "Duplicate key value" for e in exceptions)


def test_exceptions_cap_per_type_and_keep_other_types():
    """Many empty-key rows must not crowd out later email errors."""
    from services.excel_service import MAX_EXCEPTIONS_PER_TYPE, MAX_STORED_EXCEPTIONS

    rows = [[None, f"user{i}@ok.com"] for i in range(12)]
    rows += [[f"id-{i}", "not-an-email"] for i in range(8)]
    data = _xlsx(["CustomerID", "Email"], rows)
    _, stats, exceptions = run_validation(
        data,
        [_cfg("CustomerID", True), {**_cfg("Email"), "flag_email": True}],
    )
    empty_rows = {e["row_number"] for e in exceptions if e["error_type"] == "Key value is empty"}
    email_rows = {e["row_number"] for e in exceptions if e["error_type"] == "Invalid email format"}
    assert len(empty_rows) == MAX_EXCEPTIONS_PER_TYPE
    assert len(email_rows) == MAX_EXCEPTIONS_PER_TYPE
    assert len(exceptions) <= MAX_STORED_EXCEPTIONS
    assert stats["total_errors"] > len(exceptions)
